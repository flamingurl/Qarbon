import openpyxl
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os

class ExcelHandler:
    def __init__(self, workers_file='data/workers.xlsx', tasks_file='data/tasks.xlsx'):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.workers_file = os.path.join(base_dir, 'data', 'workers.xlsx')
        self.tasks_file = os.path.join(base_dir, 'data', 'tasks.xlsx')
        self._ensure_files_exist()

    def _get_est_time(self):
        # Render servers are UTC; this offsets to EST (UTC-5)
        return (datetime.utcnow() - timedelta(hours=5)).strftime('%m/%d/%Y %I:%M %p')

    def _ensure_files_exist(self):
        os.makedirs(os.path.dirname(self.workers_file), exist_ok=True)
        if not os.path.exists(self.workers_file):
            wb = Workbook()
            ws = wb.active
            ws.append(['Name', 'Job Title', 'Date Working'])
            wb.save(self.workers_file)
        if not os.path.exists(self.tasks_file):
            wb = Workbook()
            ws = wb.active
            ws.append(['Urgency', 'Task Description', 'Date Assigned', 'Date Completed', 'Assigned To'])
            wb.save(self.tasks_file)

    def read_workers(self):
        wb = load_workbook(self.workers_file)
        return [{'name': r[0], 'job_title': r[1], 'date_working': r[2]} for r in wb.active.iter_rows(min_row=2, values_only=True) if r[0]]

    def read_tasks(self):
        wb = load_workbook(self.tasks_file)
        return [{'row_number': i, 'urgency': r[0], 'description': r[1], 'date_assigned': r[2], 'date_completed': r[3], 'assigned_to': r[4]} 
                for i, r in enumerate(wb.active.iter_rows(min_row=2, values_only=True), start=2) if r[1]]

    def update_task_completion(self, row_number):
        wb = load_workbook(self.tasks_file)
        wb.active.cell(row=int(row_number), column=4, value=self._get_est_time())
        wb.save(self.tasks_file)

    def assign_task_to_worker(self, row_number, worker_name):
        wb = load_workbook(self.tasks_file)
        wb.active.cell(row=int(row_number), column=5, value=worker_name)
        wb.save(self.tasks_file)

    def delete_task(self, row_number):
        wb = load_workbook(self.tasks_file)
        wb.active.delete_rows(int(row_number))
        wb.save(self.tasks_file)

    def delete_worker(self, name):
        wb = load_workbook(self.workers_file)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == name:
                ws.delete_rows(row)
                break
        wb.save(self.workers_file)

    def add_worker(self, name, job_title, date_working):
        wb = load_workbook(self.workers_file)
        wb.active.append([name, job_title, date_working])
        wb.save(self.workers_file)

    def add_task(self, urgency, description):
        wb = load_workbook(self.tasks_file)
        est_date = (datetime.utcnow() - timedelta(hours=5)).strftime('%m/%d/%Y')
        wb.active.append([urgency, description, est_date, '', ''])
        wb.save(self.tasks_file)
